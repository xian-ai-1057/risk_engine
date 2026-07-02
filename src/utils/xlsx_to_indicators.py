"""將 Excel 指標檔轉換為 indicator.json + narrative_filter.json。

Excel 須包含兩個工作表：
  - 指標 (Sheet1)：與既有 CSV 相同欄位 — 產業別、財務分析指標、
                   指標名稱、指標對應財報欄位、指標編號、
                   指標判斷門檻值、風險情境、結果單位、敘事代碼（選用）
  - 敘事指標 (Sheet2)：
      必填欄位：產業別、段落、會計科目、會計科目代碼
      選填欄位：公式、顯示名稱、單位、替換單位
        (留白時 expression=會計科目代碼、display_name=會計科目、unit="")
        替換單位優先於單位；兩者皆留白時 unit 退回首個 code
        的「單位」（在 narrative 模組中處理）。

輸出：
  - indicator.json：{產業: [rule, ...]}（與 convert_indicators.py 相容）
  - narrative_filter.json：{產業: {段落: [
      {key, display_name, expression, unit}, ...]}}
    其中 key 在同段落內保證唯一（衝突時 append _2、_3…）。

用法:
    python -m utils.xlsx_to_indicators 指標.xlsx \\
        --config-out outputs/indicator.json \\
        --filter-out outputs/narrative_filter.json \\
        [--indicator-sheet 指標] \\
        [--filter-sheet 敘事指標]

實作備註：本模組原本透過 pandas 讀取 xlsx，pandas 會被 PyInstaller
冷凍進 EXE 帶來 ~25MB 額外體積。改用 openpyxl 直接 iter_rows
後 EXE 縮小到 ~10MB；Sheet 由 (columns, rows) tuple 表示，rows
是 list[dict[str, str]]，cell 一律轉成 stripped 字串，None / NaN 一
律 ""。
"""
import json
import logging
import sys
from typing import Any

from openpyxl import load_workbook

from utils.convert_indicators import row_to_rule

logger = logging.getLogger(__name__)


# Sheet 內部表示：(columns, rows)
# - columns：表頭順序（list[str]）
# - rows：每筆資料一個 dict，key 取自 columns，value 均為 stripped str
Sheet = tuple[list[str], list[dict[str, str]]]


# 預設 sheet 名稱與 fallback
_INDICATOR_SHEET_DEFAULT = "指標"
_FILTER_SHEET_DEFAULT = "敘事"
_TAG_SHEET_DEFAULT = "tag_table"
_INDICATOR_SHEET_FALLBACK = "Sheet1"
_FILTER_SHEET_FALLBACK = "Sheet2"

# Sheet 1 欄位（與 CSV 一致）
_INDICATOR_COLUMNS = [
    "產業別", "財務分析指標", "指標名稱",
    "指標對應財報欄位", "指標編號",
    "指標判斷門檻值", "風險情境",
]

# Sheet 2 必填欄位
_FILTER_COLUMNS = [
    "產業別", "段落", "會計科目", "會計科目代碼",
]

# Sheet 2 選填欄位（缺欄時退化為 fallback 行為）
# 每個語意欄位允許多個別名，依序取首個非空欄位的值。
# 「計算公式」是新版命名；「公式」保留向後相容。
# 「計算公式（中文表示）」屬於人類可讀說明，不參與運算，故排除在外。
# 「替換單位」優先於「單位」：若公式末端 *operand 改變了量綱
# （例如 ((A+B+C)/D)*D 結果為 仟元 而非 %），可在「替換單位」
# 直接覆寫；留白時退回「單位」/ 首個 code 的單位。
_FILTER_FORMULA_ALIASES = ("計算公式", "公式")
_FILTER_DISPLAY_NAME_ALIASES = ("顯示名稱",)
_FILTER_UNIT_ALIASES = ("替換單位", "單位")
# 「備註」欄用於標記父子關係：值為 ``{父項會計科目}子項目`` 時，
# 視為該段落內該父項的子項；其他內容視為純說明、忽略。
_FILTER_NOTE_ALIASES = ("備註",)
_PARENT_NOTE_SUFFIX = "子項目"


def _cell_str(v: Any) -> str:
    """openpyxl cell 值 → stripped str；None 一律 ""。

    float 整數值（例如 150.0）轉成 "150"，貼合 pandas dtype=str
    對 xlsx 數值 cell 的處理慣例，避免使用者輸入的整數型門檻被
    解析成 "150.0" 而破壞 threshold 解析。
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        # bool 是 int 的 subclass，得先擋掉避免被 float 分支吃到
        return str(v).strip()
    if isinstance(v, float):
        if v != v:  # NaN
            return ""
        if v.is_integer():
            return str(int(v))
    return str(v).strip()


def _first_nonempty(
    row_dict: dict[str, str],
    aliases: tuple[str, ...],
) -> str:
    """依序找第一個有值的別名欄位。"""
    for key in aliases:
        val = row_dict.get(key, "")
        if val:
            return val
    return ""


def _load_sheet(xlsx_path: str, sheet_name: str) -> Sheet:
    """讀單一 sheet → (columns, rows)。找不到 sheet 時拋 KeyError。

    cell 一律 stripped str；None / NaN → ""。每筆 row 都會補齊所有
    columns key（缺值填 ""），讓下游不必判 KeyError。
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(sheet_name)
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return [], []
        columns = [_cell_str(c) for c in header]
        # 過濾尾端空欄（pandas 不會出現，openpyxl 對某些檔案可能多出空 cell）
        while columns and columns[-1] == "":
            columns.pop()

        rows: list[dict[str, str]] = []
        for raw in rows_iter:
            values = list(raw)
            row_dict: dict[str, str] = {col: "" for col in columns}
            for idx, col in enumerate(columns):
                if idx < len(values):
                    row_dict[col] = _cell_str(values[idx])
            # 整列皆空 → 略過（避免尾端空白列被解析成資料列）
            if not any(row_dict.values()):
                continue
            rows.append(row_dict)
        return columns, rows
    finally:
        wb.close()


def _read_sheet(
    xlsx_path: str,
    primary: str,
    fallback: str,
) -> Sheet:
    """讀取指定 sheet，找不到時 fallback。"""
    try:
        return _load_sheet(xlsx_path, primary)
    except KeyError:
        logger.warning(
            "找不到工作表 '%s'，改用 '%s'",
            primary, fallback,
        )
        return _load_sheet(xlsx_path, fallback)


_RULE_REQUIRED_FIELDS = (
    "tag_id", "indicator_code", "value_formula",
)


def _validate_rule(
    rule: dict[str, Any],
    row_idx: int,
) -> list[str]:
    """檢查單筆 rule 必要欄位是否非空。

    回傳錯誤訊息列表（空 = 通過）。``row_idx`` 為 Excel 對應列號
    （1-based、含表頭時記得 +2 對齊使用者視角）。
    """
    errors: list[str] = []
    for field in _RULE_REQUIRED_FIELDS:
        val = rule.get(field, "")
        if not (val or isinstance(val, (int, float))):
            errors.append(
                f"列 {row_idx}: rule 必要欄位 '{field}' 為空 "
                f"(tag_id={rule.get('tag_id', '?')!r})"
            )
    return errors


def parse_indicator_sheet(
    sheet: Sheet,
) -> dict[str, list[dict]]:
    """Sheet 1 → {產業: [rule, ...]}。

    解析後會檢查每筆 rule 的 ``tag_id`` / ``indicator_code`` /
    ``value_formula`` 是否非空；任一筆失敗會收齊所有錯誤再
    一次拋 ValueError，方便使用者一次修完 xlsx。

    Args:
        sheet: (columns, rows) tuple，見模組頂部 ``Sheet`` 註解。

    Returns:
        與 convert_indicators.convert() 相同的結構。
    """
    columns, rows = sheet
    missing = [c for c in _INDICATOR_COLUMNS if c not in columns]
    if missing:
        raise ValueError(
            f"指標工作表缺少欄位: {', '.join(missing)}"
        )

    config: dict[str, list[dict]] = {}
    errors: list[str] = []
    for idx, row_dict in enumerate(rows):
        if not row_dict.get("產業別"):
            continue
        industries, rule = row_to_rule(row_dict)
        # Excel 列號：rows 0-based + 表頭 1 列 → +2
        excel_row = idx + 2
        rule_errors = _validate_rule(rule, excel_row)
        if rule_errors:
            errors.extend(rule_errors)
            continue
        for ind in industries:
            config.setdefault(ind, []).append(
                rule.copy(),
            )

    if errors:
        joined = "\n  ".join(errors)
        raise ValueError(
            "指標 xlsx 有 "
            f"{len(errors)} 筆規則欄位不完整：\n  {joined}"
        )

    return config


def _make_unique_key(
    base: str,
    existing_keys: set[str],
) -> str:
    """段落內 key 衝突時 append _2、_3…，回傳唯一 key。"""
    if base not in existing_keys:
        return base
    i = 2
    while f"{base}_{i}" in existing_keys:
        i += 1
    return f"{base}_{i}"


def parse_filter_sheet(
    sheet: Sheet,
) -> dict[str, dict[str, list[dict[str, Any]]]]:
    """Sheet 2 → {產業: {段落: [{key, display_name,
        expression, unit, [parent_key]}, ...]}}。

    必填欄位：產業別、段落、會計科目、會計科目代碼。
    選填欄位：公式、顯示名稱、單位、替換單位、備註
        （留白時套 fallback）。

    Fallback 規則（S1.3）：
      - expression 留白 → 會計科目代碼
      - display_name 留白 → 會計科目
      - unit：替換單位 優先；其次 單位；皆留白 → ""
        （narrative 模組階段再 fallback 至首 code 的單位）
      - key 預設 = 會計科目代碼；同段落內衝突時 append _2、_3…

    去重（S1.4）：同 (產業, 段落) 內，
        若 (會計科目代碼, expression) 完全相同則略過。

    父子層級（備註欄）：備註值符合 ``{父項會計科目}子項目`` 時，
    於同 (產業, 段落) 內以「會計科目」精確比對父項，把父項的
    ``key`` 寫進子項 entry 的 ``parent_key`` 欄。pattern 不符
    → 視為純說明、忽略；符合但找不到父項 → ``logger.warning``
    並退化為平項（不寫 ``parent_key``）。

    Args:
        sheet: (columns, rows) tuple，見模組頂部 ``Sheet`` 註解。

    Returns:
        敘事過濾結構。
    """
    columns, rows = sheet
    missing = [c for c in _FILTER_COLUMNS if c not in columns]
    if missing:
        raise ValueError(
            f"敘事指標工作表缺少欄位: {', '.join(missing)}"
        )

    result: dict[
        str, dict[str, list[dict[str, Any]]]
    ] = {}
    # (industry, section) -> set of (code, expression) 用於去重
    seen: dict[tuple[str, str], set[tuple[str, str]]] = {}
    # (industry, section) -> set of keys 用於 key 衝突檢查
    keys_per_section: dict[
        tuple[str, str], set[str]
    ] = {}
    # (industry, section) -> {display_name: key}
    # 第二階段用來解析 parent_key；同 display_name 在同段落只取
    # 首個出現的 key（理論上不應重複）。
    display_name_index: dict[
        tuple[str, str], dict[str, str]
    ] = {}
    # 等待解析父項的子項清單：(entry_ref, ind, section,
    # 父項 display_name)。第二階段查表後 mutate entry。
    pending_children: list[
        tuple[dict[str, Any], str, str, str]
    ] = []

    for row_dict in rows:
        industries_raw = row_dict.get("產業別", "")
        section = row_dict.get("段落", "")
        name = row_dict.get("會計科目", "")
        code = row_dict.get("會計科目代碼", "")
        formula_raw = _first_nonempty(
            row_dict, _FILTER_FORMULA_ALIASES,
        )
        display_name_raw = _first_nonempty(
            row_dict, _FILTER_DISPLAY_NAME_ALIASES,
        )
        unit_raw = _first_nonempty(
            row_dict, _FILTER_UNIT_ALIASES,
        )
        abs_flag_raw = (
            row_dict.get("顯示為絕對值", "") or ""
        ).strip().lower()
        note_raw = _first_nonempty(
            row_dict, _FILTER_NOTE_ALIASES,
        )

        if not (industries_raw and section and code):
            continue

        expression = formula_raw or code
        display_name = display_name_raw or name
        unit = unit_raw
        display_absolute = abs_flag_raw in (
            "是", "y", "true", "1",
        )

        # 解析備註：若以「子項目」結尾，擷取父項 display_name
        parent_display_name = ""
        if note_raw.endswith(_PARENT_NOTE_SUFFIX):
            stripped = note_raw[
                : -len(_PARENT_NOTE_SUFFIX)
            ].strip()
            if stripped:
                parent_display_name = stripped
            else:
                logger.debug(
                    "備註 '%s' 僅含『子項目』後綴、無父項名稱，"
                    "視為純說明（段落: %s, code: %s）",
                    note_raw, section, code,
                )

        industries = [
            ind.strip()
            for ind in industries_raw.split("\n")
            if ind.strip()
        ]

        for ind in industries:
            ind_bucket = result.setdefault(ind, {})
            sec_bucket = ind_bucket.setdefault(
                section, [],
            )
            sec_seen = seen.setdefault((ind, section), set())
            sec_keys = keys_per_section.setdefault(
                (ind, section), set(),
            )
            name_idx = display_name_index.setdefault(
                (ind, section), {},
            )

            dedup_key = (code, expression)
            if dedup_key in sec_seen:
                continue
            sec_seen.add(dedup_key)

            unique_key = _make_unique_key(code, sec_keys)
            sec_keys.add(unique_key)
            # 用首次出現的 key 註冊 display_name，後續查表可由
            # 子項的備註比對父項
            name_idx.setdefault(display_name, unique_key)

            entry: dict[str, Any] = {
                "key": unique_key,
                "display_name": display_name,
                "expression": expression,
                "unit": unit,
            }
            if display_absolute:
                entry["display_absolute"] = True
            sec_bucket.append(entry)

            if parent_display_name:
                pending_children.append(
                    (entry, ind, section, parent_display_name),
                )

    # 第二階段：解析 parent_key
    for entry, ind, section, parent_name in pending_children:
        parent_key = display_name_index.get(
            (ind, section), {},
        ).get(parent_name)
        if parent_key is None:
            logger.warning(
                "找不到父項『%s』（產業: %s, 段落: %s, 子項 key: %s），"
                "退化為平項",
                parent_name, ind, section, entry["key"],
            )
            continue
        if parent_key == entry["key"]:
            logger.warning(
                "備註指向自身為父項（產業: %s, 段落: %s, key: %s），"
                "忽略 parent_key",
                ind, section, entry["key"],
            )
            continue
        entry["parent_key"] = parent_key

    return result


def parse_tag_table_sheet(
    xlsx_path: str,
    sheet: str = _TAG_SHEET_DEFAULT,
) -> dict[str, str]:
    """讀 xlsx 中的 tag_table sheet → {FA_RFNBR: FA_CANME}。

    sheet 不存在時回傳空 dict（視為選用，呼叫端可 fallback 至
    EXE 同層的 ``tag_table.csv``）；欄位不齊則拋 ``ValueError``。
    """
    try:
        columns, rows = _load_sheet(xlsx_path, sheet)
    except KeyError:
        logger.info(
            "xlsx 無 '%s' sheet，跳過 tag_table 讀取", sheet,
        )
        return {}

    required = {"FA_RFNBR", "FA_CANME"}
    missing = required - set(columns)
    if missing:
        raise ValueError(
            f"tag_table sheet 缺少欄位: {', '.join(sorted(missing))}"
        )

    tag_map: dict[str, str] = {}
    for row_dict in rows:
        code = row_dict.get("FA_RFNBR", "")
        name = row_dict.get("FA_CANME", "")
        if code and name:
            tag_map[code] = name
    return tag_map


def convert(
    xlsx_path: str,
    indicator_sheet: str = _INDICATOR_SHEET_DEFAULT,
    filter_sheet: str = _FILTER_SHEET_DEFAULT,
    tag_sheet: str = _TAG_SHEET_DEFAULT,
) -> tuple[
    dict[str, list[dict]],
    dict[str, dict[str, list[dict[str, str]]]],
    dict[str, str],
]:
    """讀取 Excel，回傳 (指標 config, 敘事 filter, tag_map)。

    ``tag_map`` 為 ``{FA_RFNBR: FA_CANME}``；xlsx 中無 ``tag_table``
    sheet 時為空 dict，呼叫端可 fallback 至外部 CSV。
    """
    indicator_data = _read_sheet(
        xlsx_path, indicator_sheet,
        _INDICATOR_SHEET_FALLBACK,
    )
    filter_data = _read_sheet(
        xlsx_path, filter_sheet,
        _FILTER_SHEET_FALLBACK,
    )

    config = parse_indicator_sheet(indicator_data)
    narrative_filter = parse_filter_sheet(filter_data)
    tag_map = parse_tag_table_sheet(xlsx_path, tag_sheet)
    return config, narrative_filter, tag_map


# ── CLI ──────────────────────────────────────────────

def _parse_args(argv: list[str]) -> dict[str, Any]:
    args: dict[str, Any] = {
        "xlsx": "",
        "config_out": "indicator.json",
        "filter_out": "narrative_filter.json",
        "indicator_sheet": _INDICATOR_SHEET_DEFAULT,
        "filter_sheet": _FILTER_SHEET_DEFAULT,
    }
    flag_map = {
        "--config-out": "config_out",
        "--filter-out": "filter_out",
        "--indicator-sheet": "indicator_sheet",
        "--filter-sheet": "filter_sheet",
    }
    i = 1
    while i < len(argv):
        flag = argv[i]
        if flag in flag_map and i + 1 < len(argv):
            args[flag_map[flag]] = argv[i + 1]
            i += 2
            continue
        if not flag.startswith("--") and not args["xlsx"]:
            args["xlsx"] = flag
        i += 1
    return args


def _usage() -> None:
    print(
        "Usage: python -m utils.xlsx_to_indicators "
        "<xlsx> \\",
    )
    print("  [--config-out indicator.json] \\")
    print("  [--filter-out narrative_filter.json] \\")
    print("  [--indicator-sheet 指標] \\")
    print("  [--filter-sheet 敘事指標]")


def _write_json(path: str, data: Any) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(
            data, f, ensure_ascii=False, indent=2,
        )


def main() -> None:
    from risk_engine import log_config
    log_config.setup_logging()

    args = _parse_args(sys.argv)
    if not args["xlsx"]:
        _usage()
        sys.exit(1)

    try:
        config, narrative_filter, tag_map = convert(
            args["xlsx"],
            indicator_sheet=args["indicator_sheet"],
            filter_sheet=args["filter_sheet"],
        )
    except (FileNotFoundError, ValueError) as e:
        logger.error("轉換失敗: %s", e)
        sys.exit(1)

    _write_json(args["config_out"], config)
    _write_json(args["filter_out"], narrative_filter)

    for ind, rules in config.items():
        print(f"  指標 [{ind}]: {len(rules)} 條規則")
    for ind, sections in narrative_filter.items():
        total = sum(len(v) for v in sections.values())
        print(
            f"  敘事 [{ind}]: {len(sections)} 段落, "
            f"{total} 科目",
        )
    if tag_map:
        print(f"  tag_table: {len(tag_map)} 筆對照")

    print(f"\n已輸出指標設定至 {args['config_out']}")
    print(f"已輸出敘事過濾至 {args['filter_out']}")


if __name__ == "__main__":
    main()
